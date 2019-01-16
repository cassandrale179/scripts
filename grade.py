
from os import listdir
from os.path import isfile, join 
from openpyxl import load_workbook 

# wb = Workbook()
 
students = []



mypath = '/Users/cassandrale/Downloads/Jan8_attendance'
files = [f for f in listdir(mypath) if isfile(join(mypath, f)) and '.xlsx' in f] 


for file in files:
    i = 4
    wb = load_workbook(filename = mypath + '/' + file) 
    final_scores = wb['Final Scores'] 

    while True: 
        cell = 'B' + str(i) 
        student_name = final_scores[cell].value 
        if student_name == None:
            break; 
        else: 
            students.append(student_name)
            i += 1 
        

students.sort()
for student in students:
    print(student)
